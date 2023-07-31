from tkinter import *
import tkinter as tk
from customtkinter import *
import customtkinter as MD
from PIL import ImageTk, Image
import time
from tkinter import ttk
from tkvideo import tkvideo
from openpyxl import load_workbook
from openpyxl import *
from tkinter import messagebox
from functools import partial



def SYSTEM():
    CRMS = Tk()
    CRMS.geometry("1200x700")
    CRMS.config(bg="#000000")


    book = load_workbook("CRMSDB.xlsx")

    data = book["MAIN DATA"]




    CRMS_LOGO = Image.open("photos/LINES.png")
    CRMS_LOGO = CRMS_LOGO.resize((50,50), Image.ADAPTIVE)
    CRMS_LOGO = ImageTk.PhotoImage(CRMS_LOGO)



    system_upframe = Frame(CRMS, width=1200, height=50, bg="#545454")
    system_upframe.place(x=0, y=0)

    system_logo_label = Label(CRMS, image=CRMS_LOGO, bg="#545454")
    system_logo_label.place(x=10, y=0)



    system_uplabel = Label(system_upframe, text="CRIME RECORD MANAGEMENT SYSTEM", bg="#545454", fg="#D9D9D9",font=("Agency FB", 20, "bold"))
    system_uplabel.place(x=60, y=4)



    SEARCH_FRAME = MD.CTkFrame(CRMS, width=725, height=40, bg_color="#000000",fg_color="#A6A6A6",corner_radius=8)
    SEARCH_FRAME.place(x=30, y=100)

    search_box = MD.CTkEntry(SEARCH_FRAME,fg_color="#141414", text_color="White", font=("Agency FB", 20, "bold"))
    search_box.place(x=530,y=4,relwidth=0.23,relheight=0.82)

    #IMPORTING THE IMAGE

    search_icon = MD.CTkImage(light_image = Image.open("photos/search_icon.png"), size=(15,15))

    search_button = MD.CTkButton(SEARCH_FRAME,image=search_icon, text=None, width=5,height=10,fg_color="#141414", bg_color="#141414",hover_color="Black", command=lambda:find())
    search_button.place(x=648, y=8,relwidth=0.06)

    CASE_FRAME = MD.CTkScrollableFrame(CRMS, width=700, height=500, bg_color="#000000",fg_color="#A6A6A6",corner_radius=8)
    CASE_FRAME.place(x=30, y=150)

    GRAPH_FRAME = MD.CTkFrame(CRMS, width=380, height=515, bg_color="#000000",fg_color="#A6A6A6",corner_radius=8)
    GRAPH_FRAME.place(x=800, y=150)



    CASE_CANVAS = Canvas(CASE_FRAME, width="680", height=1200, bg="#A6A6A6")
    CASE_CANVAS.grid_propagate(False)
    CASE_CANVAS.grid(row=0, column=0)

    add_logo = Image.open("photos/add_icon.png")
    add_logo = add_logo.resize((70,70), Image.LANCZOS)
    add_logo = ImageTk.PhotoImage(add_logo)


    add_plus_case = MD.CTkButton(CRMS, text="+", font=("Ariel", 20, "bold"),width=70, height=70, hover_color="Green",fg_color="#545454", bg_color="#545454",corner_radius=0, command=lambda:add_case())
    add_plus_case.place(x=600, y=550)

    def add_case():
        CASE_FILES = Frame(CRMS, width=400, height=500, bg="#545454")
        CASE_FILES.place(x=200, y=150)

        close = MD.CTkButton(CASE_FILES, text="X", width=1, font=("Ariel", 20, "bold"),fg_color="Red",hover="Black",height=1,command=lambda:CASE_FILES.place_forget())
        close.place(x=380, y=0)

        last = data.max_row + 1
        case_number = MD.CTkLabel(CASE_FILES, text=f"CRIME CASE #{last}", font=("Agency FB", 38, "bold"), bg_color="#545454", fg_color="#D9D9D9", corner_radius=7)
        case_number.place(x=10, y=10)

        suspect_name = MD.CTkLabel(CASE_FILES, text="SUSPECT:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        suspect_name.place(x=10, y=60)



        victims_name = MD.CTkLabel(CASE_FILES, text="VICTIM:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        victims_name.place(x=10, y=100)

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=150)   

        crime_type = MD.CTkLabel(CASE_FILES, text="TYPE OF CRIMES:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        crime_type.place(x=10, y=160)

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=260)

        occurence_location = MD.CTkLabel(CASE_FILES, text="LOCATION OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_location.place(x=10, y=290)

        occurence_date = MD.CTkLabel(CASE_FILES, text="DATE OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_date.place(x=10, y=320)
    
        narrative_report = MD.CTkLabel(CASE_FILES, text="NARRATIVE REPORT:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        narrative_report.place(x=10, y=350)

        box = MD.CTkTextbox(CASE_FILES, width=360, height=70,state=NORMAL)
        box.place(x=15, y=380)

        create_s_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
        create_s_name.place(x=110, y=60, relheight=0.07)

        create_v_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
        create_v_name.place(x=110, y=100,relheight=0.07)

        
        crimes = ["Murder","Manslaughter","Assault","Robbery","Kidnapping","Rape","Domestic Violence","Burgarly","Theft","Larceny","Arson","Fraud","Identity theft","Hacking","Online Fraud"]
        create_crimes = StringVar()
        crimes_widget = MD.CTkOptionMenu(CASE_FILES,values=crimes, variable=create_crimes, fg_color="#545454",font=("Agency FB", 30, "bold"),corner_radius=5, button_color="Black", dropdown_fg_color="White", dropdown_font=("Agency FB", 20, "bold"))
        crimes_widget.place(x=15, y=210,relheight=0.07)

        create_location = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
        create_location.place(x=190, y=290,relheight=0.07)

        create_date = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
        create_date.place(x=190, y=320, relheight=0.07)

        create_button = MD.CTkButton(CASE_FILES, text="CREATE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:create())
        create_button.place(x=15, y=455)
        create_button.bind("<Button-1>", refresh)
        #the function "refresh" is binded into the button in this way, we can execute multiple functions or command in one button by binding an event


        cancel_button = MD.CTkButton(CASE_FILES, text="CANCEL", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5,command=lambda:CASE_FILES.place_forget())
        cancel_button.place(x=235, y=455)

        def create():
            if len(create_s_name.get()) == 0 or len(create_v_name.get() == 0 or len(create_crimes))==0 or len(create_location.get())==0 or len(create_date.get())==0 or len(box.get(1.0,END)==0):
                messagebox.showerror("ERROR","FILL ALL FIELDS")
            
            else:
                values_to_create = [last, create_s_name.get(), create_v_name.get(), create_crimes.get(), create_location.get(), create_date.get()]
                columns = ["A","B","C","D","E","F"]
                for i in range(len(values_to_create)):
                    data[columns[i]+str(last)].value = values_to_create[i]

                narrative_info = box.get(1.0, END)
                data["G"+str(last)].value = narrative_info

                book.save("CRMSDB.xlsx")
                messagebox.showinfo("ADDED","CREATED SUCCESSFULLY")


    def show_case(number):
        cell = number + 1
        CASE_FILES = Frame(CRMS, width=400, height=500, bg="#545454")
        CASE_FILES.place(x=200, y=150)

        serial_number = data["A"+str(number+1)].value
        case_number = MD.CTkLabel(CASE_FILES, text=f"CRIME CASE #{serial_number}", font=("Agency FB", 38, "bold"), bg_color="#545454", fg_color="#D9D9D9", corner_radius=7)
        case_number.place(x=10, y=10)


        suspect_name = MD.CTkLabel(CASE_FILES, text="SUSPECT:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        suspect_name.place(x=10, y=60)
    
        name_here = MD.CTkLabel(CASE_FILES, text=data["B"+str(cell)].value, font=("Agency FB", 30, "bold"), bg_color="#545454", corner_radius=7)
        name_here.place(x=110,y=60)

        victims_name = MD.CTkLabel(CASE_FILES, text="VICTIM:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        victims_name.place(x=10, y=100)

        v_name_here = MD.CTkLabel(CASE_FILES, text=data["C"+str(cell)].value, font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        v_name_here.place(x=110, y=100)

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=150)

        crime_type = MD.CTkLabel(CASE_FILES, text="TYPE OF CRIMES:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        crime_type.place(x=10, y=160)

        crimes = MD.CTkLabel(CASE_FILES, text=data["D"+str(cell)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        crimes.place(x=15, y=210)
    

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=260)


        occurence_location = MD.CTkLabel(CASE_FILES, text="LOCATION OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_location.place(x=10, y=290)

        location_here = MD.CTkLabel(CASE_FILES, text=data["E"+str(cell)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        location_here.place(x=190, y=290)

        occurence_date = MD.CTkLabel(CASE_FILES, text="DATE OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_date.place(x=10, y=320)


        data_here = MD.CTkLabel(CASE_FILES, text=data["F"+str(cell)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        data_here.place(x=190, y=320)

        narrative_report = MD.CTkLabel(CASE_FILES, text="NARRATIVE REPORT:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        narrative_report.place(x=10, y=350)

        box = MD.CTkTextbox(CASE_FILES, width=360, height=70,state=DISABLED)
        box.place(x=15, y=380)


        report = data["G"+str(cell)].value

        box.configure(state=NORMAL)
        box.insert(1.0, report)
        box.configure(state=DISABLED)



        edit_button = MD.CTkButton(CASE_FILES, text="EDIT", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:edit_case())
        edit_button.place(x=15, y=455)

        delete_button = MD.CTkButton(CASE_FILES, text="DELETE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:delete_record())
        delete_button.place(x=235, y=455)

        close = MD.CTkButton(CASE_FILES, text="X", width=1, font=("Ariel", 20, "bold"),fg_color="Red",hover="Black",height=1,command=lambda:CASE_FILES.place_forget())
        close.place(x=380, y=0)




        def edit_case():
            name_here.place_forget()
            v_name_here.place_forget()
            crimes.place_forget()
            location_here.place_forget()
            data_here.place_forget()
            data_here.place_forget()
            box.configure(state=NORMAL)
            box.delete(1.0, END)
            edit_button.place_forget()
            delete_button.place_forget()

            edit_s_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
            edit_s_name.place(x=110, y=60, relheight=0.07)

            edit_v_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
            edit_v_name.place(x=110, y=100,relheight=0.07)

            crimes_type = ["Murder","Manslaughter","Assault","Robbery","Kidnapping","Rape","Domestic Violence","Burgarly","Theft","Larceny","Arson","Fraud","Identity theft","Hacking","Online Fraud"]
            edit_crimes = StringVar()
            crimes_widget = MD.CTkOptionMenu(CASE_FILES, values=crimes_type, variable=edit_crimes, fg_color="#545454",font=("Agency FB", 30, "bold"),corner_radius=5, button_color="Black", dropdown_fg_color="White", dropdown_font=("Agency FB", 20, "bold"))
            crimes_widget.place(x=15, y=210,relheight=0.07)


            edit_location = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
            edit_location.place(x=190, y=290,relheight=0.07)

            edit_date = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
            edit_date.place(x=190, y=320, relheight=0.07)


            save_button = MD.CTkButton(CASE_FILES, text="SAVE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5,command=lambda:save())
            save_button.place(x=15, y=455)

            cancel_button = MD.CTkButton(CASE_FILES, text="CANCEL", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:cancel())
            cancel_button.place(x=235, y=455)


            edit_s_name.insert(0, data["B"+str(cell)].value)
            edit_v_name.insert(0,data["C"+str(cell)].value)
            edit_location.insert(0,data["E"+str(cell)].value)
            edit_date.insert(0,data["F"+str(cell)].value)
            box.insert(1.0, data["G"+str(cell)].value)



            def save():    
                if len(edit_s_name.get()) == 0 or len(edit_v_name.get()==0 or len(edit_crimes.get())==0 or len(edit_location.get())==0 or len(edit_date.get())==0):
                    messagebox.showerror("ERROR","FILL ALL FIELDS")

                else:
                    values_to_replace=[edit_s_name.get(), edit_v_name.get(), edit_crimes.get(), edit_location.get(), edit_date.get()]
                    columns = ["B","C","D","E","F"]
                    for i in range(len(values_to_replace)):
                        data[columns[i]+str(cell)].value = values_to_replace[i]            
                    
                    narrative_info = box.get(1.0, END)
                    data["G"+str(cell)].value = narrative_info

                    book.save("CRMSDB.xlsx")
                    messagebox.showinfo("DONE","SAVED SUCCESSFULLY")



            def cancel():
                edit_s_name.place_forget()
                edit_v_name.place_forget()
                crimes_widget.place_forget()
                edit_location.place_forget()
                edit_date.place_forget()
                save_button.place_forget()
                cancel_button.place_forget()



                name_here.place(x=110,y=60)
                v_name_here.place(x=110, y=100)
                crimes.place(x=15, y=210)
                location_here.place(x=190, y=290)
                data_here.place(x=190, y=320)
                report = data["G"+str(cell)].value
                box.configure(state=NORMAL)
                box.insert(1.0, report)
                box.configure(state=DISABLED)
                edit_button.place(x=15, y=455)
                delete_button.place(x=235, y=455)


        def delete_record():
            if number > 1:
                data.delete_rows(number+1)
                book.save("CRMSDB.xlsx")
                messagebox.showinfo("DONE","RECORD DELETED SUCCESSFULLY")
                btns[-1].grid_forget()
                btns.pop()


            else:
                print("you're trying to delete the column names")



    #length is the numbers of cases inside the database
    length = data.max_row - 1

    indicator = 0

    own_logo = Image.open("photos/case_logo.png")
    own_logo = own_logo.resize((115,115), Image.LANCZOS)
    own_logo = ImageTk.PhotoImage(own_logo)

    #the parameter is divided by 2 since we do not indiscriminately place the widget, since we
    #a grid we need to have a griding system where as we limit the column to four and 
    #and the the number of column inside the row is determine by the algorithm below



    btns = []
    a = 0
    for row in range(length):
        if indicator != length:
            for column in range(4):
                a+=1
                l = MD.CTkFrame(CASE_CANVAS, width=10, height=1)
                #l = Button(CASE_CANVAS, text=f"CASE NUMBER# {a}", command=partial(show_case, a))
                s = MD.CTkButton(l, text=None,image=own_logo, width=10, height=10,fg_color="#545454", hover_color="Red", command=partial(show_case,a))
                s.pack()
                
                case_names = MD.CTkLabel(l, text="CRIME CASE", font=("Agency FB", 20, "bold"))
                case_names.pack(side=BOTTOM)


                l.grid(row=row, column=column, padx=20, pady=20)
                btns.append(l)
                indicator += 1

                print(f"row={row}    column={column}")
                if indicator == length:
                    break

        else:
            break
        
    print(a)



    #refreshes the content of the frame
    def refresh(event):
        length = data.max_row - 1
        indicator = 0
        btns = []
        a = 0
        for row in range(length):
            if indicator != length:
                for column in range(4):
                    a+=1
                    l = MD.CTkFrame(CASE_CANVAS, width=10, height=1)
                    #l = Button(CASE_CANVAS, text=f"CASE NUMBER# {a}", command=partial(show_case, a))
                    s = MD.CTkButton(l, text=None,image=own_logo, width=10, height=10,fg_color="#545454", hover_color="Red", command=partial(show_case,a))
                    s.pack()
                    
                    case_names = MD.CTkLabel(l, text="CRIME CASE", font=("Agency FB", 20, "bold"))
                    case_names.pack(side=BOTTOM)


                    l.grid(row=row, column=column, padx=20, pady=20)
                    btns.append(l)
                    indicator += 1

                    print(f"row={row}    column={column}")
                    if indicator == length:
                        break

            else:
                break

    def find():
        if len(search_box.get())==0:
            messagebox.showerror("ERROR","FILL THE FIELDS")
        else:
            value_to_find = search_box.get()
            

            found = "Hindi"
            for i in range(2, data.max_row):
                if search_box.get() == str(data["A"+str(i-1)].value):
                    found = "NAHANAP"
                else:
                    1+1
            if found == "NAHANAP":
                messagebox.showinfo("FOUND","FOUND IT")
            else:
                messagebox.showerror("NOT FOUND","DATA DOESNT EXIST")
                    
    CRMS.mainloop()


SYSTEM()