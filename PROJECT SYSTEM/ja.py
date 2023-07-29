from cgitb import text
from email.mime import image
from enum import Flag
from itertools import tee
from textwrap import fill
from tkinter import *
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from turtle import bgcolor, width
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import PIL
from PIL import Image, ImageTk

root = tk.Tk()
root.geometry('400x300')
root.title('WELCOME DANCERS')
root.configure(bg="#7fc3c0")
#root.resizable(width=FALSE, height= FALSE)

#1 establish a connection
excel_con = Workbook()

#2 Get the actual file
excel_con = load_workbook('finalproject.xlsx')

#3 activate 
excel_activate = excel_con.active

def login():
    root.withdraw()
    new()

def new():
    global newpage
    newpage = tk.Toplevel() 
    newpage.geometry('1000x700')
    newpage.title("Next Dimension")
    newpage.configure(bg="#141414")

    def update():
        change = tk.Toplevel()
        change.geometry('300x100')
        change.title('SEARCH')
        change.configure(bg="#98DBC6")

        fname = Label(change, text="NAME:",bg=("#98DBC6"),font=("Times New Roman",15))
        fname.grid(row=0, column=0, padx=10)
        findthis = Entry(change, font=("Times New Roman",15))
        findthis.grid(row=0, column=1, columnspan=2)
        btn = Button(change, text="FIND", command=lambda:searchandchange(), font=("Times New Roman",10), width=10)
        btn.grid(row=1, column=1, columnspan=2, pady=10)

        def searchandchange():
            scc = tk.Toplevel()
            scc.geometry('500x300')
            scc.title('UPDATE')
            scc.configure(bg="#98DBC6")
            valv = findthis.get()

            Found = False
            tango = 0
            
            for i in range(2, excel_activate.max_row):
                if valv == excel_activate["A"+str(i)].value:
                    Found = True
                    tango = i
                    break
                else:
                    continue
            
            if Found == True:
                messagebox.showerror("Error", "DATA MUST EXIST TO UPDATE")
            else:
                findthis.grid_forget()
                btn.grid_forget()

                a = Entry(scc, width=50)
                b = Entry(scc, width=50)
                c = Entry(scc, width=50)
                d = StringVar()
                a1 = Label(scc,text="NAME", bg="#98DBC6")
                a2 = Label(scc,text="EMAIL", bg="#98DBC6")
                a3 = Label(scc,text="PHONE NUMBER", bg="#98DBC6")
                a4 = Label(scc,text="GENDER", bg="#98DBC6")
                a5 = Label(scc,text="ADDRESS", bg="#98DBC6")
                a6 = Label(scc,text="TYPE OF DANCE", bg="#98DBC6")
            
                male_button = ttk.Radiobutton(scc, text="MALE", variable= d, value="MALE")
                female_button = ttk.Radiobutton(scc, text="FEMALE", variable= d, value="FEMALE")
                e = Text(scc, height=5, width=20)

                f = StringVar()
                dance_list = ["SALSA", "WALTZ", "JIVE", "HIP-HOP", "STREET DANCE"]
                dance_combo = ttk.Combobox(scc, values= dance_list, textvariable= f, width=40) 

                a.grid(row=0, column=1, columnspan=2)
                b.grid(row=1, column=1, columnspan=2)
                c.grid(row=2, column=1, columnspan=2)

                a1.grid(row=0, column=0)
                a2.grid(row=1, column=0)
                a3.grid(row=2, column=0)
                a4.grid (row=3, column=0)
                a5.grid(row=4, column=0)
                a6.grid(row=5, column=0)

                male_button.grid(row=3, column=1)
                female_button.grid(row=3, column=2)
                e.grid(row=4, column=1, columnspan=2)
                dance_combo.grid(row=5, column=1, columnspan=3)


                changeIT = Button(scc, text="UPDATE", command=lambda:overwrite())
                changeIT.grid(row=6, column=1, pady=20)

                def overwrite():
                    va = a.get()
                    vb = b.get()
                    vc = c.get()
                    vd = d.get()
                    ve = e.get(0.1)

                    excel_activate["A"+str(tango)].value = va
                    excel_activate["B"+str(tango)].value = vb
                    excel_activate["C"+str(tango)].value = vc
                    excel_activate["D"+str(tango)].value = vd
                    excel_activate["E"+str(tango)].value = ve

                    excel_con.save("finalproject.xlsx")               
                    messagebox.showinfo("SAVED","DATA SUCCESSFULLY UPDATED")

    def save():
        a = name1.get()    
        b = email1.get()
        c = pn.get()
        d = gen_b.get()
        e = address1.get(0.1, END)
        f = tod1.get()
        
        result = True

        for i in range (2, excel_activate.max_row + 1):
            if a == excel_activate['A' + str(i)].value or b == excel_activate['B' + str(i)].value or c == excel_activate['C' + str(i)].value or f == excel_activate['F' + str(i)].value:
                    result = False
            break
        if result == False:
            messagebox.showerror("DATA", "DATA ALREADY EXIST")
        else:
            lastrow = str(excel_activate.max_row +1)
            excel_activate['A'+ lastrow] = a
            excel_activate['B'+ lastrow] = b
            excel_activate['C'+ lastrow] = c
            excel_activate['D'+ lastrow] = d
            excel_activate['E'+ lastrow] = e
            excel_activate['F'+ lastrow] = f
                       
            excel_con.save('finalproject.xlsx')
            messagebox.showinfo("SAVED", "DATA SAVED")
     
    def search ():
            searchb = tk.Toplevel()
            searchb.geometry('300x100')
            searchb.title('SEARCH')
            searchb.configure(bg="#98DBC6")

          
            sname = Label (searchb, text="NAME",bg= ("#98DBC6"), font=("Times New Roman",15))
            sname.grid(row=0, column=0)

            
            namevalue = Entry(searchb, font=("Times New Roman",15))
            namevalue.grid(row=0, column=1, padx=10, columnspan=2)

            btn1 = Button(searchb, text="Search",command=lambda:find(), font=("Times New Roman",10), width=10)
            btn1.grid(row=2, column=1, pady=15)

            btn2 = Button(searchb, text="Exit", command=lambda:searchb.withdraw(), font=("Times New Roman",10), width=10)
            btn2.grid(row=2, column=2, pady=15)

            def find():
                names = namevalue.get()

                Found = False
                position = 0

                for i in range(2, excel_activate.max_row):
                    if names == excel_activate["A"+str(i)].value:
                        Found = True
                        position += i
                        break
                    else:
                        continue
                
                if Found == True:
                    messagebox.showerror("ERROR","DATA NOT FOUND")
                else:
                    messagebox.showinfo("FOUND", f"{names} IS FOUND AT cell address row {position}")
                
    def view_data():
    # Create a new top-level window
        vdata = tk.Toplevel()
        vdata.title("Data Viewer")
        vdata.geometry("500x500")
        vdata.resizable(width=False, height=False)

        tree = ttk.Treeview(vdata)
        treescrolly = Scrollbar(tree, orient="vertical", command=tree.yview)
        treescrollx = Scrollbar(tree, orient="horizontal", command=tree.xview)
        tree.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
        treescrollx.pack(side ="bottom",fill ="x")
        treescrolly.pack(side ="right",fill="y")

        tree["columns"] = ("NAME", "MAIL","PHONENUM", "GENDER","ADDRESS", "DANCETYPE")

        tree.column("#0", width=0, stretch=False)
        tree.column("NAME", width=120)
        tree.column("MAIL", width=120)
        tree.column("PHONENUM", width=120)
        tree.column("GENDER", width=120)
        tree.column("ADDRESS", width=120)
        tree.column("DANCETYPE", width=120)

        tree.heading("#0", text=None)
        tree.heading("NAME", text="NAME")
        tree.heading("MAIL", text="MAIL")
        tree.heading("PHONENUM", text="PHONENUM")
        tree.heading("GENDER", text="GENDER")
        tree.heading("ADDRESS", text="ADDRESS")
        tree.heading("DANCETYPE", text="DANCETYPE")

        for i in range(2, excel_activate.max_row + 1):
            tree.insert(parent="", index=i, iid=i, values=(excel_activate["A"+str(i)].value,
                                                        excel_activate["B"+str(i)].value,
                                                        excel_activate["C"+str(i)].value,
                                                        excel_activate["D"+str(i)].value,
                                                        excel_activate["E"+str(i)].value,
                                                        excel_activate["F"+str(i)].value))
        tree.pack(side="left", fill="both", expand=True)

    def delete ():
        d = Toplevel()
        d.title("DELETE")
        d.configure(bg="#98DBC6")
        d.geometry("300x100")

        l = Label (d, text="NAME",bg="#98DBC6", font=("Times New Roman",15))
        l.grid(row=0,column=0)

        la = Entry(d, font=("Times New Roman",15))
        la.grid(row=0, column=1, columnspan=2, padx=10)

        lab = ttk.Button(d, text="DELETE", width=10, command=lambda:ddelete())
        lab.grid(row=2, column=1, pady=15)

        btnD = ttk.Button(d, text="EXIT", command=lambda:d.withdraw(), width=10)
        btnD.grid(row=2, column=2, pady=15)

        def ddelete():
            for every_cell in range(2, (excel_activate.max_row)+1):
                if (la.get() == excel_activate['A'+str( every_cell)].value):
                
                    Found = True
                    cell_address = every_cell
                
                    break
                else:
                    Found=False
            if(Found == True):
                excel_activate.delete_rows(cell_address)
            messagebox.showerror("DELETE","DATA DELETED")
            excel_con.save('finalproject.xlsx')
            clear_entries()

        def clear_entries():
            name1.delete(0,END)
            email1.delete(0,END)
            pn.delete(0,END)
            gen_b.set("")
            address1.delete(0.1,END)
            tod1.set("")
    
    topframe= Frame (newpage, height=80, width=1000, bg="#7fc3c0",highlightbackground="#7fc3c0", highlightthickness=3)
    topframe.grid(row=0, column=0, columnspan=2, pady=10)
    topframe.grid_propagate(False)

    reg = Label (topframe, text=" REGISTRATION FOR DANCE WORKSHOP", font=("Times New Roman",30), bg="#000000",fg="White")
    reg.pack(fill='x', expand=True)

    leftframe = Frame(newpage, height=500, width=800, bg="#cfb845",highlightbackground="#7fc3c0", highlightthickness=3)
    leftframe.grid(row=1, column=0,columnspan=2, pady=15)
    leftframe.grid_propagate(False)

    buttomframe= Frame (newpage, height=90, width=1000, bg="#141414",highlightbackground="#141414", highlightthickness=3)
    buttomframe.grid(row=2, column= 0, columnspan=2, pady=10)
    buttomframe.grid_propagate(False) 

    #Label
    name = Label(leftframe, text = "FULL NAME", font=("Times New Roman",20), bg="#cfb845")
    name.grid(row=0, column=0, padx= 5, pady=15 ,sticky=W)
    email = Label(leftframe, text = "EMAIL:", font=("Times New Roman",20), bg="#cfb845")
    email.grid(row=1, column=0, padx= 5, pady=15 ,sticky=W)
    pnum = Label(leftframe, text = "PHONE NUMBER:", font=("Times New Roman",20), bg="#cfb845")
    pnum.grid(row=2, column=0, padx= 5, pady=15 ,sticky=W)
    gender = Label(leftframe, text = "GENDER:", font=("Times New Roman",20), bg="#cfb845")
    gender.grid(row=3, column=0, padx= 5, pady=15 ,sticky=W)
    address =Label(leftframe, text = "ADDRESS:", font=("Times New Roman",20), bg="#cfb845")
    address.grid(row=4, column=0, padx= 5, pady=15,sticky=W)
    tod = Label(leftframe, text = "TYPE OF DANCE:", font=("Times New Roman",20), bg="#cfb845")
    tod.grid(row=5, column=0, padx= 5, pady=15 ,sticky=W)

    #Entry
    global name1
    global email1
    global pn
    global gen_b
    global address1
    global tod1
  
    name1 = Entry (leftframe, width=40, font=("Times New Roman",15))
    name1.grid(row=0, column=2, columnspan=2)
    email1 = Entry (leftframe, width=40, font=("Times New Roman",15))
    email1.grid(row=1, column=2, columnspan=2)
    pn= Entry (leftframe, width=40, font=("Times New Roman",15))
    pn.grid(row=2, column=2, columnspan=2)

    gen_b = StringVar()
    male_button = ttk.Radiobutton(leftframe, text="MALE", variable= gen_b, value="MALE")
    female_button = ttk.Radiobutton(leftframe, text="FEMALE", variable= gen_b, value="FEMALE")
    male_button.grid(row=3, column= 2, pady=10)
    female_button.grid(row=3, column= 3, pady=10)

    address1 = Text (leftframe, height=5, width=40, font=("Times New Roman",15))
    address1.grid(row=4, column=2, columnspan=2)

    tod1 = StringVar()
    dance_list = ["SALSA", "WALTZ", "JIVE", "HIP-HOP", "STREET DANCE"]
    dance_combo = ttk.Combobox(leftframe, values= dance_list, textvariable= tod1, width=38, font=("Times New Roman",15))  
    dance_combo.grid(row=5, column=2, columnspan=2, padx=80)

    #Button
    b1 = Button(buttomframe, text="UPDATE", font=("Times New Roman",15),bg= "LightBlue", command=lambda:update())
    b1.grid(row=1, column=0, padx=(50,20), pady =10)
    b2 = Button(buttomframe, text="SAVE", font=("Times New Roman",15),bg= "LightBlue", command=lambda:save())
    b2.grid(row=1, column=1, padx=(50,20), pady = 10)
    b3 = Button(buttomframe, text="SEARCH", font=("Times New Roman",15),bg= "LightBlue", command=lambda:search())
    b3.grid(row=1, column=2, padx=(50,20), pady = 10)
    b4 = Button(buttomframe, text="DELETE", font=("Times New Roman",15),bg= "LightBlue", command=lambda:delete())
    b4.grid(row=1, column=3, padx=(50,20), pady = 10)
    b6 = Button(buttomframe, text="VIEW DATA", font=("Times New Roman",15),bg= "LightBlue",command=lambda:view_data())
    b6.grid(row=1, column=4, padx=(50,20), pady = 10)
    b5 = Button(buttomframe, text="EXIT", font=("Times New Roman",15),bg= "LightBlue",command=lambda:root.destroy())
    b5.grid(row=1, column=5, padx=(50,20), pady = 10)

    newpage.mainloop()

# #LABEL

bgs = Image.open("reg.png")
bgs = bgs.resize((400,300), Image.LANCZOS)
bgs = ImageTk.PhotoImage(bgs)

title = Label(root, image=bgs, font=("Ariel", 13))
title.place(x=0, y=0)

login_b = Button(root, text="REGISTER", bg="#7fc3c0",command=lambda:login())

login_b.place(x=170, y=250)

root.mainloop()