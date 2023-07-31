from tkinter import *
import tkinter as tk
from customtkinter import *
import customtkinter as MD
from PIL import ImageTk, Image
import time
from tkinter import ttk
from tkvideo import tkvideo
from openpyxl import load_workbook
from openpyxl import *
from tkinter import messagebox
from functools import partial


#LOADINNG ANIMATION
origin = Tk()
origin.overrideredirect(1)
#size of window and coordinates
taas = 500
wide = 300
xcor = 750
ycor = 250

origin.geometry(f"{taas}x{wide}+{xcor}+{ycor}")
origin.update_idletasks()

origin.config(bg="#141414")

#-----------------------------------------------------------------FRONT END-------------------------------------------------

btn = Image.open("photos/logo.png")
btn = btn.resize((150, 150), Image.LANCZOS)
btn = ImageTk.PhotoImage(btn)


systemlogo = Label(origin, image=btn, bg="#141414")
systemlogo.place(x=170, y=20)

title = Label(origin, text="C R M S", fg="#004FFF", bg="#141414", font=("Valorant", 20))
title.place(x=193, y=160)

dot1 = Image.open("photos/loaddots/4.png")
dot1 = dot1.resize((20,20), Image.LANCZOS)
dot1 = ImageTk.PhotoImage(dot1)

dot2 = Image.open("photos/loaddots/3.png")
dot2 = dot2.resize((20,20), Image.LANCZOS)
dot2 = ImageTk.PhotoImage(dot2)

dot3 = Image.open("photos/loaddots/2.png")
dot3 = dot3.resize((20,20), Image.LANCZOS)
dot3 = ImageTk.PhotoImage(dot3)

dot4 = Image.open("photos/loaddots/1.png")
dot4 = dot4.resize((20,20), Image.LANCZOS)
dot4 = ImageTk.PhotoImage(dot4)


dotmate1 = Label(origin, image=dot1, bg="#141414")
dotmate1.place(x=192, y=200)

dotmate2 = Label(origin, image=dot2, bg="#141414")
dotmate2.place(x=222, y=200)

dotmate3 = Label(origin, image=dot3, bg="#141414")
dotmate3.place(x=252, y=200)

dotmate4 = Label(origin, image=dot4, bg="#141414")
dotmate4.place(x=282, y=200)


origin.update_idletasks()
startnotice = Label(origin, text="Starting...", bg="#141414", fg="White", font=((23)))
startnotice.place(x=0, y=275)
time.sleep(0.5)


start = 0
while start < 5:
    dotmate1.config(image=dot4)
    dotmate2.config(image=dot1)
    dotmate3.config(image=dot2)
    dotmate4.config(image=dot3)
    time.sleep(0.5)
    origin.update_idletasks()

    dotmate1.config(image=dot3)
    dotmate2.config(image=dot4)
    dotmate3.config(image=dot1)
    dotmate4.config(image=dot2)
    time.sleep(0.5)
    origin.update_idletasks()

    dotmate1.config(image=dot2)
    dotmate2.config(image=dot3)
    dotmate3.config(image=dot4)
    dotmate4.config(image=dot1)
    time.sleep(0.5)
    origin.update_idletasks()

    dotmate1.config(image=dot1)
    dotmate2.config(image=dot2)
    dotmate3.config(image=dot3)
    dotmate4.config(image=dot4)
    time.sleep(0.5)
    origin.update_idletasks()
    start += 1




origin.withdraw()
LogSign = Toplevel()
LogSign.geometry("1200x700")
#LogSign.overrideredirect(1)
LogSign.config(bg="#141414")


#LOGO
uplogo = Image.open("photos\main_in_logo.png")
uplogo = uplogo.resize((50,50), Image.LANCZOS)
uplogo = ImageTk.PhotoImage(uplogo)



upframe = Frame(LogSign, width=1200, height=50, bg="#545454")
upframe.place(x=0, y=0)

mainlogo = Label(upframe, image=uplogo, bg="#545454")
mainlogo.place(x=10, y=0)

uplabel = Label(upframe, text="CRIME RECORD MANAGEMENT SYSTEM", bg="#545454", fg="#D9D9D9",font=("Agency FB", 20, "bold"))
uplabel.place(x=60, y=4)

uplogintbn = MD.CTkButton(upframe, text="LOGIN", width=5,fg_color="#545454",font=("Agency FB", 30, "bold"), command=lambda:slideleft())
uplogintbn.place(relx=0.85, y=4)

upsignupbtn = MD.CTkButton(upframe, text="SIGNUP", width=5,fg_color="#545454",font=("Agency FB", 30, "bold"), command=lambda:sign_slide_left())
upsignupbtn.place(relx=0.92, y=4)


#frame to contain the video
vidframe = Frame(LogSign, width=1200, height=650,bg="#545454")
vidframe.place(x=-4, y=50)

a = Label(vidframe,bg="#545454")
a.pack()

video = tkvideo("intro_vid.mp4", a, loop=True, size=(1209, 650))
video.play()


a = Label(LogSign, text="INDICATOR")
a.place(x=1200, y=500)









x_axis = 1300
login = MD.CTkFrame(LogSign, height=400, width=350, bg_color="#141414",fg_color="#D9D9D9",corner_radius=0)
login.place(x=x_axis, y=70)


a = Label(login, text="USER LOGIN", font=("Cambria", 20), bg="#D9D9D9")
a.place(relx=0.280, rely=0.05)

b = Entry(login, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
b.place(relx=0.12, rely=0.3, relheight=0.08)

c = Entry(login, width=37, bd=0, font=("Ariel", 10, "bold"), show="*",bg="#D9D9D9")
c.place(relx=0.12, rely=0.5, relheight=0.08)


e = MD.CTkButton(login, text="LOG IN", font=("Corbel", 16,"bold"), width=272, hover_color="Red", fg_color="#61777B",bg_color="White",corner_radius=0,command=lambda:gate())
e.place(relx=0.1, rely=0.7, relheight=0.08)

f = MD.CTkButton(login, text="Cancel", font=("Corbel", 16), width=272, hover_color="Red", bg_color="White",corner_radius=0,command=lambda:slideright())
f.place(relx=0.1, rely=0.8, relheight=0.08)




linesimg = Image.open("photos/linya.PNG")
lineimg = linesimg.resize((270,2), Image.LANCZOS)
lineimg = ImageTk.PhotoImage(lineimg)


ID = Label(login, image=lineimg, bg="Black")
ID.place(relx=0.1, rely=0.375, relheight=0.005)

keycode = Label(login, image=lineimg, bg="Black")
keycode.place(relx=0.1, rely=0.575, relheight=0.005)



#CREATING THE ANIMATION FOR USER ID LABEL
def moveup():
    global ycor
    if ycor != 100:
        ycor -=1
        user.place(x=35, y=ycor)
        if ycor > 100:
            login.after(5, moveup)
        else:
                print("FLOWN")
    


def moveback():
    global ycor
    if ycor != 123:
        ycor += 1
        user.place(x=35, y=ycor)
        if ycor < 123:
            login.after(5, moveback)
        else:
            print("LANDED")

indicator = 0
def animation(event):
    global indicator
    indicator += 1
    if indicator % 2 == 0:
        if len(b.get()) > 0:
            1+1
        else:
            moveback()
    else:
        moveup()

#CREATING ANIMATION FOR PASSCODE LABEL


def go_up():
    global codecor
    if codecor != 180:
        codecor -= 1
        passcode.place(x=35, y=codecor)
        if codecor > 180:
            login.after(5, go_up)
        else:
            print("ON")

def go_down():
    global codecor
    if codecor != 203:
        codecor += 1
        passcode.place(x=35, y=codecor)
        if codecor < 203:
            login.after(5, go_down)
        else:
            print("OFF")

switch = 0
def anotheranimation(event):
    global switch
    switch += 1
    if switch % 2 == 0:
        if len(c.get()) > 0:
            1+1
        else:
            go_down()
    else:
        go_up()




#usecor = 100
ycor = 123
user = Label(login, text="USER ID",font=("Corbel", 13,"bold"), bg="#D9D9D9")
user.place(x=35, y=ycor)

#passcor = 180
codecor = 203
passcode = Label(login, text="PASSCODE",font=("Corbel", 13,"bold"), bg="#D9D9D9")
passcode.place(x=35, y=codecor)

b.bind("<Enter>", animation)
b.bind("<Leave>", animation)

c.bind("<Enter>", anotheranimation)
c.bind("<Leave>", anotheranimation)


#340
#700
#ANIMATE THE FRAME

def slideleft():
    if negCor == 500:
        print("RETRACT THE SIGNUP FRAME FIRST")
    else:
        global x_axis
        if x_axis != 820:
            x_axis -= 20
            login.place(x=x_axis, y=70)
            if x_axis > 820:
                LogSign.after(10, slideleft)

def slideright():

    global x_axis
    if x_axis != 1300:
        x_axis += 20
        login.place(x=x_axis, y=70)
        if x_axis < 1300:
            LogSign.after(10, slideright)
#_---------------------------------------------------ABOVE ANIMATION FOR FRAMES----------------------------------------



#--------------------------------------------BELOW ANIMATION FOR SIGNUP PANEL AND WIDGETS------------------------------

negCor=1200
signup = MD.CTkFrame(LogSign, width=650, height=400, bg_color="#141414",fg_color="#D9D9D9",corner_radius=0)
#hide
signup.place(x=negCor, y=70)

#show
#signup.place(x=0, y=negCor)


FRAME = MD.CTkScrollableFrame(signup, width=605, height=350, bg_color="#D9D9D9", fg_color="#D9D9D9",corner_radius=0)
FRAME.place(x=4, y=4)

CANVAS = Canvas(FRAME, width=620, height=600, highlightbackground="#D9D9D9", background="#D9D9D9")
CANVAS.pack(fill=BOTH)











signtitle = Label(CANVAS, text="USER SIGNUP", font=("Cambria", 20), bg="#D9D9D9")
signtitle.place(relx=0.340, rely=0.1)

entry_height = 0.06




#FOR X COORDINATES OF WIDGETS inside the signup panel for easy modification
row1 = 15
row2 = 325


name = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
name.place(x=row1, y=160, relheight=entry_height)

Identification = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
Identification.place(x=row2, y=160, relheight=entry_height)

email = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
email.place(x=row1, y=240, relheight=entry_height)

password = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
password.place(x=row2, y=240, relheight=entry_height)

confirmpassword = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
confirmpassword.place(x=row2, y=320, relheight=entry_height)

phonenumber = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
phonenumber.place(x=row1, y=320, relheight=entry_height)

dateofbirth = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
dateofbirth.place(x=row1, y=400, relheight=entry_height)

zipcode = Entry(CANVAS, width=37, bd=0, font=("Ariel", 10, "bold"), bg="#D9D9D9")
zipcode.place(x=row2, y=400, relheight=entry_height)

linesimg = Image.open("photos/linya.PNG")
lineimg = linesimg.resize((255,2), Image.LANCZOS)
lineimg = ImageTk.PhotoImage(lineimg)


#LABEL TO ANIMATE
#VARIABLES FOR Y COORDINATE OF EACH LABEL
firstrow = 160
secondrow = 240
thirdrow = 320
fourthrow = 400




fullname = Label(CANVAS, text="FULL NAME",font=("Corbel", 13,"bold"), bg="#D9D9D9")
fullname.place(x=row1, y=firstrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row1-1, y=190)

ID = Label(CANVAS, text="USER ID",font=("Corbel", 13,"bold"), bg="#D9D9D9")
ID.place(x=row2,y=firstrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row2-1, y=190)


mail = Label(CANVAS, text="EMAIL",font=("Corbel", 13,"bold"), bg="#D9D9D9")
mail.place(x=row1,y=secondrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row1-1, y=270)

word = Label(CANVAS, text="PASSWORD",font=("Corbel", 13,"bold"), bg="#D9D9D9")
word.place(x=row2,y=secondrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row2-1, y=270)

conword = Label(CANVAS, text="CONFIRM PASSWORD",font=("Corbel", 13,"bold"), bg="#D9D9D9")
conword.place(x=row2,y=thirdrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row2-1, y=350)

phonenum = Label(CANVAS, text="PHONE NUMBER",font=("Corbel", 13,"bold"), bg="#D9D9D9")
phonenum.place(x=row1,y=thirdrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row1-1, y=350)

bday = Label(CANVAS, text="DATE OF BIRTH",font=("Corbel", 13,"bold"), bg="#D9D9D9")
bday.place(x=row1,y=fourthrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row1-1, y=430)

Zcode = Label(CANVAS, text="ZipCode",font=("Corbel", 13,"bold"), bg="#D9D9D9")
Zcode.place(x=row2,y=fourthrow)
line = Label(CANVAS, image=lineimg)
line.place(x=row2-1, y=430)

#####################################-------------ANIMATING THE LABEL WIDGETS-------------------###########################

#-------------------------------------------------NAME LABEL ANIMATION BELOW-------------------------------------------------
#NAME LABEL
#for label NAME to go up
def nameup():
    global firstrow
    if firstrow != 142:
        firstrow-=1
        fullname.place(x=row1, y=firstrow)
        if firstrow > 142:
            signup.after(10, nameup)

#for label NAME to go down

def namedown():
    global firstrow
    if firstrow != 160:
        firstrow +=1
        fullname.place(x=row1, y=firstrow)
        if firstrow < 160:
            signup.after(10, namedown)

#ANIMATION FUNCTION OF NAME LABEL
name_indicator = 0
def name_animation(event):
    global name_indicator
    name_indicator += 1
    if (name_indicator % 2)==0:
        if len(name.get())>0:
            1+1
        else:
            namedown()
    else:
        nameup()
        
name.bind("<Enter>", name_animation)
name.bind("<Leave>", name_animation)
#-----------------------------------------------NAME LABEL ANIMATION ABOVE-------------------------------------------------




#-----------------------------------------------USER LABEL ANIMATION BELOW-------------------------------------------------

#USER ID
#for label USER ID to go up

def user_up():
    global firstrow
    if firstrow != 142:
        firstrow -= 1
        ID.place(x=row2,y=firstrow)
        if firstrow > 142:
            signup.after(10,user_up)
        else:
            print("DAP")


def user_down():
    global firstrow
    if firstrow != 160:
        firstrow += 1
        ID.place(x=row2,y=firstrow)
        if firstrow < 160:
            signup.after(10,user_down)
        else:
            print("DOP")

user_indicator = 0
def user_animation(event):
    global user_indicator
    user_indicator+=1
    if (user_indicator % 2)==0:
        if len(Identification.get())>0:
            1+1
        else:
            user_down()
    else:
        user_up()




Identification.bind("<Enter>", user_animation)
Identification.bind("<Leave>", user_animation)
#-----------------------------------------------USER LABEL ANIMATION ABOVE-------------------------------------------------

#-----------------------------------------------EMAIL LABEL ANIMATION BELOW-------------------------------------------------

#to go up
def mail_up():
    global secondrow
    if secondrow != 222:
        secondrow -= 1
        mail.place(x=row1,y=secondrow)
        if secondrow > 222:
            signup.after(10, mail_up)
#TO GO DOWN
def mail_down():
    global secondrow
    if secondrow != 240:
        secondrow += 1
        mail.place(x=row1,y=secondrow)
        if secondrow < 240:
            signup.after(10, mail_down)


mail_indicator = 0

def mail_animation(event):
    global mail_indicator
    mail_indicator += 1
    if (mail_indicator % 2) == 0:
        if len(email.get())>0:
            #DO NOTHING
            1+1
        else:
            mail_down()
    else:
        mail_up()


email.bind("<Enter>", mail_animation)
email.bind("<Leave>", mail_animation)



#-----------------------------------------------EMAIL LABEL ANIMATION ABOVE-------------------------------------------------

#-----------------------------------------------PASSWORD LABEL ANIMATION BELOW-------------------------------------------------
#TO GO UP
def pass_up():
    global secondrow
    if secondrow != 222:
        secondrow -= 1
        word.place(x=row2,y=secondrow)
        if secondrow > 222:
            signup.after(10, pass_up)

#TO GO DOWN
def pass_down():
    global secondrow
    if secondrow != 240:
        secondrow += 1
        word.place(x=row2,y=secondrow)
        if secondrow < 240:
            signup.after(10, pass_down)

password_indicator = 0
def password_animation(event):
    global password_indicator
    password_indicator += 1
    if (password_indicator % 2)==0:
        if len(password.get())>0:
            #do nothing
            1+1
        else:
            pass_down()
    else:
        pass_up()

password.bind("<Enter>", password_animation)
password.bind("<Leave>", password_animation)

#-----------------------------------------------PASSWORD LABEL ANIMATION ABOVE-------------------------------------------------

#---------------------------------------------PHONE NUMBER LABEL ANIMATION BELOW-------------------------------------------------
#TO GO UP
def phone_up():
    global thirdrow
    if thirdrow != 302:
        thirdrow -= 1
        phonenum.place(x=row1,y=thirdrow)
        if thirdrow > 302:
            signup.after(10, phone_up)
    
#TO GO DOWN
def phone_down():
    global thirdrow
    if thirdrow != 320:
        thirdrow += 1
        phonenum.place(x=row1,y=thirdrow)
        if thirdrow < 320:
            signup.after(10, phone_down)

phone_label_indicator = 0
def phone_label_animation(event):
    global phone_label_indicator
    phone_label_indicator+=1
    if (phone_label_indicator % 2) == 0:
        if len(phonenumber.get()) > 0:
            1+1
            #do nothing
        else:
            phone_down()
    else:
        phone_up()

phonenumber.bind("<Enter>", phone_label_animation)
phonenumber.bind("<Leave>", phone_label_animation)

        
#---------------------------------------------PHONE NUMBER LABEL ANIMATION ABOVE-------------------------------------------------

#----------------------------------------CONFIRM PASSWORD NUMBER LABEL ANIMATION BELOW-------------------------------------------------
#TO GO UP
def confirm_up():
    global thirdrow
    if thirdrow != 302:
        thirdrow -= 1
        conword.place(x=row2,y=thirdrow)
        if thirdrow > 302:
            signup.after(10, confirm_up)

#TO GO DOWN
def confirm_down():
    global thirdrow
    if thirdrow != 320:
        thirdrow += 1
        conword.place(x=row2,y=thirdrow)
        if thirdrow < 320:
            signup.after(10, confirm_down)


confirm_indicator = 0
def confirm_pass_animation(event):
    global confirm_indicator
    confirm_indicator += 1
    if (confirm_indicator % 2)==0:
        if len(confirmpassword.get())>0:
            #do nothing
            1+1
        else:
            confirm_down()
    else:
        confirm_up()

confirmpassword.bind("<Enter>", confirm_pass_animation)
confirmpassword.bind("<Leave>", confirm_pass_animation)


#----------------------------------------CONFIRM PASSWORD NUMBER LABEL ANIMATION ABOVE-------------------------------------------------

#----------------------------------------BIRTHDAY LABEL ANIMATION BELOW-------------------------------------------------
#TO GO UP
def date_up():
    global fourthrow
    if fourthrow != 382:
        fourthrow -= 1
        bday.place(x=row1,y=fourthrow)
        if fourthrow > 382:
            signup.after(10, date_up)

#TO GO DOWN
def date_down():
    global fourthrow
    if fourthrow != 400:
        fourthrow += 1
        bday.place(x=row1,y=fourthrow)
        if fourthrow < 400:
            signup.after(10, date_down)

date_indicator = 0
def date_animation(event):
    global date_indicator
    date_indicator+=1
    if (date_indicator%2)==0:
        if len(dateofbirth.get())>0:
            #do nothing
            1+1
        else:
            date_down()
    else:
        date_up()

dateofbirth.bind("<Enter>", date_animation)
dateofbirth.bind("<Leave>", date_animation)



#----------------------------------------BIRTHDAY LABEL LABEL ANIMATION ABOVE-------------------------------------------------

#----------------------------------------ZIPCODE LABEL LABEL ANIMATION BELOW-------------------------------------------------

def code_up():
    global fourthrow
    if fourthrow != 382:
        fourthrow -= 1
        Zcode.place(x=row2,y=fourthrow)
        if fourthrow > 382:
            signup.after(10, code_up)

def code_down():
    global fourthrow
    if fourthrow != 400:
        fourthrow += 1
        Zcode.place(x=row2,y=fourthrow)
        if fourthrow < 400:
            signup.after(10, code_down)

Zindicator = 0
def Zcode_animation(event):
    global Zindicator
    Zindicator += 1
    if (Zindicator % 2)==0:
        if len(zipcode.get())>0:
            #do nothing
            1+1
        else:
            code_down()
    else:
        code_up()

zipcode.bind("<Enter>", Zcode_animation)
zipcode.bind("<Leave>", Zcode_animation)

#ANIMATE THE FRAME CONTAINER
def sign_slide_left():
    if x_axis == 820:
        print("RETRACT THE LOGIN FRAME FIRST")
    else:
        global negCor
        if negCor != 500:
            negCor -= 20
            signup.place(x=negCor, y=70)
            if negCor > 500:
                LogSign.after(10, sign_slide_left)


def sign_slide_right():
    global negCor
    if negCor != 1200:
        negCor +=20
        signup.place(x=negCor, y=70)
        if negCor < 1200:
            LogSign.after(10, sign_slide_right)


#----------------------------------------ZIPCODE LABEL LABEL ANIMATION ABOVE-------------------------------------------------





#BUTTONS
register = MD.CTkButton(CANVAS, text="REGISTER", font=("Agency FB", 30, "bold"),text_color="Black",fg_color="#61777B",corner_radius=0,command=lambda:put_in_db())
register.place(x=row1, y=450, relwidth=0.41,relheight=entry_height)

cancelregister = MD.CTkButton(CANVAS, text="CANCEL", font=("Agency FB", 30, "bold"),text_color="Black",fg_color="#61777B",corner_radius=0,command=lambda:sign_slide_right())
cancelregister.place(x=row2, y=450, relwidth=0.41,relheight=entry_height)








#--------------------------------------------ABOVE ANIMATION FOR SIGNUP PANEL AND WIDGETS------------------------------



#LOADING THE EXCEL IN THE PROGRAM
book = load_workbook("CRMSDB.xlsx")

#ASSIGNING THE SPREADSHEET OF CRMSDB.XLSX
LOG_IN = book["CREDENTIALS"]
data = book["MAIN DATA"]



######################LOGGING IN
def gate():
    user_id = b.get()
    pass_code = c.get()
    global LOGGED_ON_USER

    if len(user_id) <= 0 or len(pass_code) <= 0:
        messagebox.showerror("FIELD","FILL ALL FIELDS TO LOGIN")
    else:
        found = False
        last_cell = LOG_IN.max_row
        LOGGED_ON_USER = None
        for i in range(2, last_cell+1):
            if LOG_IN["A"+str(i)].value == user_id and LOG_IN["B"+str(i)].value == pass_code:
                found = True
                LOGGED_ON_USER = i
                break
                
            else:
                found = False

        if found == True:
            #calls the system interface which is inside of function
            SYSTEM()
        else:
            messagebox.showwarning("LOGIN","WRONG CREDENTIALS")







######################LOGGING IN

######################SIGNUP

def put_in_db():
    sign_name = name.get()
    sign_user_id = Identification.get()
    sign_email = email.get()
    sign_password = password.get()
    sign_confirm_pass = confirmpassword.get()
    sign_phonenumber = phonenumber.get()
    sign_bday = dateofbirth.get()
    sign_zip = zipcode.get()

    



    if len(sign_name) <= 0 or len(sign_user_id) <= 0 or len(sign_email) <= 0 or len(sign_password) <= 0 or len(sign_confirm_pass) <= 0 or len(sign_bday) <= 0 or len(sign_zip) <= 0:
        messagebox.showerror("404","FILL ALL FIELDS TO PROCEED")
    else:
        if sign_password != sign_confirm_pass:
            messagebox.showerror("202","PASSWORD DO NOT MATCH")
        else:
            found = "HINDI"
            for i in range(2,LOG_IN.max_row+1):
                if sign_user_id == LOG_IN["A"+str(i)].value or sign_name == LOG_IN["C"+str(i)].value:
                    found = "NAHANAP"
                    
                
            if found == "NAHANAP":
                messagebox.showwarning("LOGIN","ACCOUNT ALREADY EXIST")
            else:
                okay_to_go = False
                try:
                    int(sign_phonenumber)
                    okay_to_go = True
                except ValueError:
                    okay_to_go = False

                if okay_to_go == False:
                    messagebox.showerror("ERROR","PHONE NUMBER SHOULD BE A NUMERIC VALUE")
                else:
                    last_row = str(LOG_IN.max_row+1)

                    all_value = ["#",sign_user_id, sign_password, sign_name,sign_phonenumber, sign_email, sign_bday, sign_zip]
                    all_column = ["#","A","B","C","D","E","F","G"]
                    for i in range(1,len(all_column)):
                        LOG_IN[all_column[i]+last_row].value = all_value[i]
                    
                    print("REGISTED")
                    messagebox.showinfo("SUCCESS","REGISTERED")

                    book.save("CRMSDB.xlsx")











#-----------------------------------------------------------------FRONT END-------------------------------------------------





#----------------------------------------------------------------MAIN INTERFACE----------------------------------------

def SYSTEM():
    LogSign.withdraw()

    CRMS = Toplevel()
    CRMS.geometry("1200x700")
    CRMS.config(bg="Black")





    system_upframe = Frame(CRMS, width=1200, height=50, bg="#545454")
    system_upframe.place(x=0, y=0)


    global system_logo
    system_logo = Image.open("photos/main_in_logo.png")
    system_logo = system_logo.resize((42,42), Image.LANCZOS)
    system_logo = ImageTk.PhotoImage(system_logo)

    CRMS_LOGO = Label(CRMS, image=system_logo, bg="#545454")
    CRMS_LOGO.place(x=30, y=4)




    system_uplabel = Label(system_upframe, text="CRIME RECORD MANAGEMENT SYSTEM", bg="#545454", fg="#D9D9D9",font=("Agency FB", 20, "bold"))
    system_uplabel.place(x=80, y=4)




    SEARCH_FRAME = MD.CTkFrame(CRMS, width=725, height=40, bg_color="#000000",fg_color="#A6A6A6",corner_radius=8)
    SEARCH_FRAME.place(x=30, y=100)

    search_box = MD.CTkEntry(SEARCH_FRAME,fg_color="#141414", text_color="White", font=("Agency FB", 20, "bold"))
    search_box.place(x=530,y=4,relwidth=0.23,relheight=0.82)

    #IMPORTING THE IMAGE

    search_icon = MD.CTkImage(light_image = Image.open("photos/search_icon.png"), size=(15,15))

    search_button = MD.CTkButton(SEARCH_FRAME,image=search_icon, text=None, width=5,height=10,fg_color="#141414", bg_color="#141414",hover_color="Black", command=lambda:find())
    search_button.place(x=648, y=8,relwidth=0.06)

    CASE_FRAME = MD.CTkScrollableFrame(CRMS, width=700, height=500, bg_color="#000000",fg_color="#A6A6A6",corner_radius=8)
    CASE_FRAME.place(x=30, y=150)

    INFO_FRAME = MD.CTkFrame(CRMS, width=380, height=515, bg_color="#000000",fg_color="#A6A6A6",corner_radius=8)
    INFO_FRAME.place(x=800, y=150)



    CASE_CANVAS = Canvas(CASE_FRAME, width="680", height=1200, bg="#A6A6A6")
    CASE_CANVAS.grid_propagate(False)
    CASE_CANVAS.grid(row=0, column=0)

    add_logo = Image.open("photos/add_icon.png")
    add_logo = add_logo.resize((70,70), Image.LANCZOS)
    add_logo = ImageTk.PhotoImage(add_logo)


    add_plus_case = MD.CTkButton(CRMS, text="+", font=("Ariel", 20, "bold"),width=70, height=70, hover_color="Red",fg_color="Green", bg_color="Green",corner_radius=0, command=lambda:add_case())
    add_plus_case.place(x=600, y=550)

    global avatar
    avatar = Image.open("photos/profile.png")
    avatar = avatar.resize((80,80), Image.LANCZOS)
    avatar = ImageTk.PhotoImage(avatar)

    profile_photo = MD.CTkLabel(CRMS, text=None,image=avatar)
    profile_photo.place(x=920, y=55)



    profile = MD.CTkButton(CRMS, text=LOG_IN["C"+str(LOGGED_ON_USER)].value, font=("Agency FB", 20, "bold"))
    profile.place(x=1010, y=60)

    log_out = MD.CTkButton(CRMS, text="LOG OUT",fg_color="Red",hover_color="Green",command=lambda:back(),corner_radius=5)
    log_out.place(x=1010, y=100)


    def back():
        LogSign.deiconify()
        CRMS.withdraw()



    def add_case():
        CASE_FILES = Frame(CRMS, width=400, height=500, bg="#545454")
        CASE_FILES.place(x=200, y=150)

        close = MD.CTkButton(CASE_FILES, text="X", width=1, font=("Ariel", 20, "bold"),fg_color="Red",hover="Black",height=1,command=lambda:CASE_FILES.place_forget())
        close.place(x=380, y=0)

        last = data.max_row + 1
        case_number = MD.CTkLabel(CASE_FILES, text=f"CRIME CASE #{last}", font=("Agency FB", 38, "bold"), bg_color="#545454", fg_color="#D9D9D9", corner_radius=7)
        case_number.place(x=10, y=10)

        suspect_name = MD.CTkLabel(CASE_FILES, text="SUSPECT:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        suspect_name.place(x=10, y=60)



        victims_name = MD.CTkLabel(CASE_FILES, text="VICTIM:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        victims_name.place(x=10, y=100)

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=150)   

        crime_type = MD.CTkLabel(CASE_FILES, text="TYPE OF CRIMES:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        crime_type.place(x=10, y=160)

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=260)

        occurence_location = MD.CTkLabel(CASE_FILES, text="LOCATION OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_location.place(x=10, y=290)

        occurence_date = MD.CTkLabel(CASE_FILES, text="DATE OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_date.place(x=10, y=320)
    
        narrative_report = MD.CTkLabel(CASE_FILES, text="NARRATIVE REPORT:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        narrative_report.place(x=10, y=350)

        box = MD.CTkTextbox(CASE_FILES, width=360, height=70,state=NORMAL)
        box.place(x=15, y=380)

        create_s_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
        create_s_name.place(x=110, y=60, relheight=0.07)

        create_v_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
        create_v_name.place(x=110, y=100,relheight=0.07)

        
        crimes = ["Murder","Manslaughter","Assault","Robbery","Kidnapping","Rape","Domestic Violence","Burgarly","Theft","Larceny","Arson","Fraud","Identity theft","Hacking","Online Fraud"]
        create_crimes = StringVar()
        crimes_widget = MD.CTkOptionMenu(CASE_FILES,values=crimes, variable=create_crimes, fg_color="#545454",font=("Agency FB", 30, "bold"),corner_radius=5, button_color="Black", dropdown_fg_color="White", dropdown_font=("Agency FB", 20, "bold"))
        crimes_widget.place(x=15, y=210,relheight=0.07)

        create_location = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
        create_location.place(x=190, y=290,relheight=0.07)

        create_date = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
        create_date.place(x=190, y=320, relheight=0.07)

        create_button = MD.CTkButton(CASE_FILES, text="CREATE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:create())
        create_button.place(x=15, y=455)
        create_button.bind("<Button-1>", refresh)
        #the function "refresh" is binded into the button in this way, we can execute multiple functions or command in one button by binding an event


        cancel_button = MD.CTkButton(CASE_FILES, text="CANCEL", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5,command=lambda:CASE_FILES.place_forget())
        cancel_button.place(x=235, y=455)

        def create():
            if len(create_s_name.get()) == 0 or len(create_v_name.get() == 0 or len(create_crimes))==0 or len(create_location.get())==0 or len(create_date.get())==0 or len(box.get(1.0,END)==0):
                messagebox.showerror("ERROR","FILL ALL FIELDS")
            
            else:
                values_to_create = [last, create_s_name.get(), create_v_name.get(), create_crimes.get(), create_location.get(), create_date.get()]
                columns = ["A","B","C","D","E","F"]
                for i in range(len(values_to_create)):
                    data[columns[i]+str(last)].value = values_to_create[i]

                narrative_info = box.get(1.0, END)
                data["G"+str(last)].value = narrative_info

                book.save("CRMSDB.xlsx")
                messagebox.showinfo("ADDED","CREATED SUCCESSFULLY")


    def show_case(number):
        cell = number + 1
        CASE_FILES = Frame(CRMS, width=400, height=500, bg="#545454")
        CASE_FILES.place(x=200, y=150)

        serial_number = data["A"+str(number+1)].value
        case_number = MD.CTkLabel(CASE_FILES, text=f"CRIME CASE #{serial_number}", font=("Agency FB", 38, "bold"), bg_color="#545454", fg_color="#D9D9D9", corner_radius=7)
        case_number.place(x=10, y=10)


        suspect_name = MD.CTkLabel(CASE_FILES, text="SUSPECT:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        suspect_name.place(x=10, y=60)
    
        name_here = MD.CTkLabel(CASE_FILES, text=data["B"+str(cell)].value, font=("Agency FB", 30, "bold"), bg_color="#545454", corner_radius=7)
        name_here.place(x=110,y=60)

        victims_name = MD.CTkLabel(CASE_FILES, text="VICTIM:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        victims_name.place(x=10, y=100)

        v_name_here = MD.CTkLabel(CASE_FILES, text=data["C"+str(cell)].value, font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        v_name_here.place(x=110, y=100)

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=150)

        crime_type = MD.CTkLabel(CASE_FILES, text="TYPE OF CRIMES:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        crime_type.place(x=10, y=160)

        crimes = MD.CTkLabel(CASE_FILES, text=data["D"+str(cell)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        crimes.place(x=15, y=210)
    

        lines = Image.open("photos/linya.PNG")
        lines = lines.resize((350,1), Image.LANCZOS)
        lines = ImageTk.PhotoImage(lines)

        lines = Label(CASE_FILES, image=lines)
        lines.place(x=20, y=260)


        occurence_location = MD.CTkLabel(CASE_FILES, text="LOCATION OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_location.place(x=10, y=290)

        location_here = MD.CTkLabel(CASE_FILES, text=data["E"+str(cell)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        location_here.place(x=190, y=290)

        occurence_date = MD.CTkLabel(CASE_FILES, text="DATE OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        occurence_date.place(x=10, y=320)


        data_here = MD.CTkLabel(CASE_FILES, text=data["F"+str(cell)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        data_here.place(x=190, y=320)

        narrative_report = MD.CTkLabel(CASE_FILES, text="NARRATIVE REPORT:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
        narrative_report.place(x=10, y=350)

        box = MD.CTkTextbox(CASE_FILES, width=360, height=70,state=DISABLED)
        box.place(x=15, y=380)


        report = data["G"+str(cell)].value

        box.configure(state=NORMAL)
        box.insert(1.0, report)
        box.configure(state=DISABLED)



        edit_button = MD.CTkButton(CASE_FILES, text="EDIT", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:edit_case())
        edit_button.place(x=15, y=455)

        delete_button = MD.CTkButton(CASE_FILES, text="DELETE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:delete_record())
        delete_button.place(x=235, y=455)

        close = MD.CTkButton(CASE_FILES, text="X", width=1, font=("Ariel", 20, "bold"),fg_color="Red",hover="Black",height=1,command=lambda:CASE_FILES.place_forget())
        close.place(x=380, y=0)




        def edit_case():
            name_here.place_forget()
            v_name_here.place_forget()
            crimes.place_forget()
            location_here.place_forget()
            data_here.place_forget()
            data_here.place_forget()
            box.configure(state=NORMAL)
            box.delete(1.0, END)
            edit_button.place_forget()
            delete_button.place_forget()

            edit_s_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
            edit_s_name.place(x=110, y=60, relheight=0.07)

            edit_v_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
            edit_v_name.place(x=110, y=100,relheight=0.07)

            crimes_type = ["Murder","Manslaughter","Assault","Robbery","Kidnapping","Rape","Domestic Violence","Burgarly","Theft","Larceny","Arson","Fraud","Identity theft","Hacking","Online Fraud"]
            edit_crimes = StringVar()
            crimes_widget = MD.CTkOptionMenu(CASE_FILES, values=crimes_type, variable=edit_crimes, fg_color="#545454",font=("Agency FB", 30, "bold"),corner_radius=5, button_color="Black", dropdown_fg_color="White", dropdown_font=("Agency FB", 20, "bold"))
            crimes_widget.place(x=15, y=210,relheight=0.07)


            edit_location = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
            edit_location.place(x=190, y=290,relheight=0.07)

            edit_date = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
            edit_date.place(x=190, y=320, relheight=0.07)


            save_button = MD.CTkButton(CASE_FILES, text="SAVE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5,command=lambda:save())
            save_button.place(x=15, y=455)

            cancel_button = MD.CTkButton(CASE_FILES, text="CANCEL", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:cancel())
            cancel_button.place(x=235, y=455)


            edit_s_name.insert(0, data["B"+str(cell)].value)
            edit_v_name.insert(0,data["C"+str(cell)].value)
            edit_location.insert(0,data["E"+str(cell)].value)
            edit_date.insert(0,data["F"+str(cell)].value)
            box.insert(1.0, data["G"+str(cell)].value)



            def save():    
                if len(edit_s_name.get())==0 or len(edit_v_name.get())==0 or edit_crimes.get()==None or len(edit_location.get())==0 or len(edit_date.get())==0:
                    messagebox.showerror("ERROR","FILL ALL FIELDS")

                else:
                    values_to_replace=[edit_s_name.get(), edit_v_name.get(), edit_crimes.get(), edit_location.get(), edit_date.get()]
                    columns = ["B","C","D","E","F"]
                    for i in range(len(values_to_replace)):
                        data[columns[i]+str(cell)].value = values_to_replace[i]            
                    
                    narrative_info = box.get(1.0, END)
                    data["G"+str(cell)].value = narrative_info

                    book.save("CRMSDB.xlsx")
                    messagebox.showinfo("DONE","SAVED SUCCESSFULLY")



            def cancel():
                edit_s_name.place_forget()
                edit_v_name.place_forget()
                crimes_widget.place_forget()
                edit_location.place_forget()
                edit_date.place_forget()
                save_button.place_forget()
                cancel_button.place_forget()



                name_here.place(x=110,y=60)
                v_name_here.place(x=110, y=100)
                crimes.place(x=15, y=210)
                location_here.place(x=190, y=290)
                data_here.place(x=190, y=320)
                report = data["G"+str(cell)].value
                box.configure(state=NORMAL)
                box.insert(1.0, report)
                box.configure(state=DISABLED)
                edit_button.place(x=15, y=455)
                delete_button.place(x=235, y=455)


        def delete_record():
            if number > 1:
                data.delete_rows(number+1)
                book.save("CRMSDB.xlsx")
                messagebox.showinfo("DONE","RECORD DELETED SUCCESSFULLY")
                btns[-1].grid_forget()
                btns.pop()


            else:
                print("you're trying to delete the column names")



    #length is the numbers of cases inside the database
    length = data.max_row - 1

    indicator = 0

    own_logo = Image.open("photos/case_logo.png")
    own_logo = own_logo.resize((115,115), Image.LANCZOS)
    own_logo = ImageTk.PhotoImage(own_logo)

    #the parameter is divided by 2 since we do not indiscriminately place the widget, since we use
    #a grid we need to have a griding system where as we limit the column to four and 
    #and the the number of column inside the row is determine by the algorithm below



    btns = []
    a = 0
    for row in range(length):
        if indicator != length:
            for column in range(4):
                a+=1
                l = MD.CTkFrame(CASE_CANVAS, width=10, height=1)
                #l = Button(CASE_CANVAS, text=f"CASE NUMBER# {a}", command=partial(show_case, a))
                s = MD.CTkButton(l, text=None,image=own_logo, width=10, height=10,fg_color="#545454", hover_color="Red", command=partial(show_case,a))
                s.pack()
                
                case_names = MD.CTkLabel(l, text="CRIME CASE", font=("Agency FB", 20, "bold"))
                case_names.pack(side=BOTTOM)


                l.grid(row=row, column=column, padx=20, pady=20)
                btns.append(l)
                indicator += 1

                print(f"row={row}    column={column}")
                if indicator == length:
                    break

        else:
            break
        
    print(a)



    #refreshes the content of the frame
    def refresh(event):
        length = data.max_row - 1
        indicator = 0
        btns = []
        a = 0
        for row in range(length):
            if indicator != length:
                for column in range(4):
                    a+=1
                    l = MD.CTkFrame(CASE_CANVAS, width=10, height=1)
                    #l = Button(CASE_CANVAS, text=f"CASE NUMBER# {a}", command=partial(show_case, a))
                    s = MD.CTkButton(l, text=None,image=own_logo, width=10, height=10,fg_color="#545454", hover_color="Red", command=partial(show_case,a))
                    s.pack()
                    
                    case_names = MD.CTkLabel(l, text="CRIME CASE", font=("Agency FB", 20, "bold"))
                    case_names.pack(side=BOTTOM)


                    l.grid(row=row, column=column, padx=20, pady=20)
                    btns.append(l)
                    indicator += 1

                    print(f"row={row}    column={column}")
                    if indicator == length:
                        break

            else:
                break

    def find():
        if len(search_box.get())==0:
            messagebox.showerror("ERROR","FILL THE FIELDS")
        else:
            value_to_find = search_box.get()
            

            found = "Hindi"
            position = 0
            for i in range(2, data.max_row):
                if search_box.get() == str(data["A"+str(i)].value):
                    found = "NAHANAP"
                    position = i
                else:
                    1+1
            if found == "Hindi":
                messagebox.showerror("NOT FOUND","DATA DOESNT EXIST")
            elif found == "NAHANAP":
                print("found it")
                GOT = Toplevel()
                GOT.geometry("400x500")
                CASE_FILES = Frame(GOT, width=400, height=500, bg="#545454")
                CASE_FILES.place(x=0, y=0)

                serial_number = search_box.get()
                case_number = MD.CTkLabel(CASE_FILES, text=f"CRIME CASE #{serial_number}", font=("Agency FB", 38, "bold"), bg_color="#545454", fg_color="#D9D9D9", corner_radius=7)
                case_number.place(x=10, y=10)


                suspect_name = MD.CTkLabel(CASE_FILES, text="SUSPECT:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                suspect_name.place(x=10, y=60)
            
                name_here = MD.CTkLabel(CASE_FILES, text=data["B"+str(position)].value, font=("Agency FB", 30, "bold"), bg_color="#545454", corner_radius=7)
                name_here.place(x=110,y=60)

                victims_name = MD.CTkLabel(CASE_FILES, text="VICTIM:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                victims_name.place(x=10, y=100)

                v_name_here = MD.CTkLabel(CASE_FILES, text=data["C"+str(position)].value, font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                v_name_here.place(x=110, y=100)

                lines = Image.open("photos/linya.PNG")
                lines = lines.resize((350,1), Image.LANCZOS)
                lines = ImageTk.PhotoImage(lines)

                lines = Label(CASE_FILES, image=lines)
                lines.place(x=20, y=150)

                crime_type = MD.CTkLabel(CASE_FILES, text="TYPE OF CRIMES:", font=("Agency FB", 30, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                crime_type.place(x=10, y=160)

                crimes = MD.CTkLabel(CASE_FILES, text=data["D"+str(position)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                crimes.place(x=15, y=210)
            

                lines = Image.open("photos/linya.PNG")
                lines = lines.resize((350,1), Image.LANCZOS)
                lines = ImageTk.PhotoImage(lines)

                lines = Label(CASE_FILES, image=lines)
                lines.place(x=20, y=260)


                occurence_location = MD.CTkLabel(CASE_FILES, text="LOCATION OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                occurence_location.place(x=10, y=290)

                location_here = MD.CTkLabel(CASE_FILES, text=data["E"+str(position)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                location_here.place(x=190, y=290)

                occurence_date = MD.CTkLabel(CASE_FILES, text="DATE OF OCCURENCE:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                occurence_date.place(x=10, y=320)


                data_here = MD.CTkLabel(CASE_FILES, text=data["F"+str(position)].value, font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                data_here.place(x=190, y=320)

                narrative_report = MD.CTkLabel(CASE_FILES, text="NARRATIVE REPORT:", font=("Agency FB", 20, "bold"), bg_color="#545454", fg_color="#545454", corner_radius=7)
                narrative_report.place(x=10, y=350)

                box = MD.CTkTextbox(CASE_FILES, width=360, height=70,state=DISABLED)
                box.place(x=15, y=380)


                report = data["G"+str(position)].value

                box.configure(state=NORMAL)
                box.insert(1.0, report)
                box.configure(state=DISABLED)



                edit_button = MD.CTkButton(CASE_FILES, text="EDIT", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:edit_searched_case())
                edit_button.place(x=15, y=455)

                delete_button = MD.CTkButton(CASE_FILES, text="DELETE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5)
                delete_button.place(x=235, y=455)

                close = MD.CTkButton(CASE_FILES, text="X", width=1, font=("Ariel", 20, "bold"),fg_color="Red",hover="Black",height=1,command=lambda:GOT.withdraw())
                close.place(x=380, y=0)

                def edit_searched_case():
                    name_here.place_forget()
                    v_name_here.place_forget()
                    crimes.place_forget()
                    location_here.place_forget()
                    data_here.place_forget()
                    data_here.place_forget()
                    box.configure(state=NORMAL)
                    box.delete(1.0, END)
                    edit_button.place_forget()
                    delete_button.place_forget()

                    edit_s_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
                    edit_s_name.place(x=110, y=60, relheight=0.07)

                    edit_v_name = MD.CTkEntry(CASE_FILES,width=250,font=("Agency FB", 30, "bold"))
                    edit_v_name.place(x=110, y=100,relheight=0.07)

                    crimes_type = ["Murder","Manslaughter","Assault","Robbery","Kidnapping","Rape","Domestic Violence","Burgarly","Theft","Larceny","Arson","Fraud","Identity theft","Hacking","Online Fraud"]
                    edit_crimes = StringVar()
                    crimes_widget = MD.CTkOptionMenu(CASE_FILES,values=crimes_type, variable=edit_crimes, fg_color="#545454",font=("Agency FB", 30, "bold"),corner_radius=5, button_color="Black", dropdown_fg_color="White", dropdown_font=("Agency FB", 20, "bold"))
                    crimes_widget.place(x=15, y=210,relheight=0.07)


                    edit_location = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
                    edit_location.place(x=190, y=290,relheight=0.07)

                    edit_date = MD.CTkEntry(CASE_FILES,width=200,font=("Agency FB", 30, "bold"))
                    edit_date.place(x=190, y=320, relheight=0.07)


                    save_button = MD.CTkButton(CASE_FILES, text="SAVE", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5,command=lambda:save_searched_case())
                    save_button.place(x=15, y=455)

                    cancel_button = MD.CTkButton(CASE_FILES, text="CANCEL", font=("Agency FB", 30, "bold"), text_color="Black",bg_color="#545454", fg_color="#D9D9D9", corner_radius=5, command=lambda:cancel_searched_case())
                    cancel_button.place(x=235, y=455)


                    edit_s_name.insert(0, data["B"+str(position)].value)
                    edit_v_name.insert(0,data["C"+str(position)].value)
                    edit_location.insert(0,data["E"+str(position)].value)
                    edit_date.insert(0,data["F"+str(position)].value)
                    box.insert(1.0, data["G"+str(position)].value)

                    def save_searched_case():
                        if len(edit_s_name.get())==0 or len(edit_v_name.get())==0 or edit_crimes.get()==None or len(edit_location.get())==0 or len(edit_date.get())==0:
                            messagebox.showerror("ERROR","FILL ALL FIELDS")

                        else:
                            values_to_replace=[edit_s_name.get(), edit_v_name.get(), edit_crimes.get(), edit_location.get(), edit_date.get()]
                            columns = ["B","C","D","E","F"]
                            for i in range(len(values_to_replace)):
                                data[columns[i]+str(position)].value = values_to_replace[i]            
                            
                            narrative_info = box.get(1.0, END)
                            data["G"+str(position)].value = narrative_info

                            book.save("CRMSDB.xlsx")
                            messagebox.showinfo("DONE","SAVED SUCCESSFULLY")




                    def cancel_searched_case():
                        edit_s_name.place_forget()
                        edit_v_name.place_forget()
                        crimes_widget.place_forget()
                        edit_location.place_forget()
                        edit_date.place_forget()
                        save_button.place_forget()
                        cancel_button.place_forget()



                        name_here.place(x=110,y=60)
                        v_name_here.place(x=110, y=100)
                        crimes.place(x=15, y=210)
                        location_here.place(x=190, y=290)
                        data_here.place(x=190, y=320)
                        edit_button.place(x=15, y=455)
                        delete_button.place(x=235, y=455)


    global info_frame_logo
    info_frame_logo = Image.open("photos/info_logo.png")
    info_frame_logo = info_frame_logo.resize((60,60), Image.LANCZOS)
    info_frame_logo = ImageTk.PhotoImage(info_frame_logo)


    PHOTO = Label(INFO_FRAME, image=info_frame_logo,bg="#A6A6A6")
    PHOTO.place(x=25, y=30)

    info_system = Label(INFO_FRAME, text="CRIME RECORD MANAGEMENT SYSTEM", bg="#A6A6A6", fg="Black",font=("Agency FB", 16, "bold"))
    info_system.place(x=95, y=45)


    sentence = f"""

    The crime record management 
    system is a sophisticated software
    tool designed to effectively organize
    and maintain crime records,
    facilitating law enforcement agencies
    in their efforts to combat 
    crime, analyze trends, and ensure 
    public safety.

    """

    introduction = Label(INFO_FRAME, text=sentence, height=10,font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    introduction.place(x=7, y=120)
    

    profile_photo = MD.CTkLabel(CRMS, text=None,image=avatar)
    profile_photo.place(x=920, y=55)


    title = Label(INFO_FRAME, text="PROFILE",font=("Agency FB", 30, "bold"),bg="#A6A6A6")

    show_name = Label(INFO_FRAME, text="NAME:",font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    show_id = Label(INFO_FRAME, text="USER ID:",font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    show_email = Label(INFO_FRAME, text="EMAIL:",font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    show_phone_number = Label(INFO_FRAME, text="PHONE NUMBER:", font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    
    data_show_name = Label(INFO_FRAME, text=LOG_IN["C"+str(LOGGED_ON_USER)].value,font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    data_show_id = Label(INFO_FRAME, text=LOG_IN["A"+str(LOGGED_ON_USER)].value,font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    data_show_email = Label(INFO_FRAME, text=LOG_IN["E"+str(LOGGED_ON_USER)].value,font=("Agency FB", 20, "bold"),bg="#A6A6A6")
    data_show_phone = Label(INFO_FRAME, text=LOG_IN["D"+str(LOGGED_ON_USER)].value,font=("Agency FB", 20, "bold"),bg="#A6A6A6")


    def user_info(event):
        introduction.place_forget()

        title.place(x=130, y=80)

        show_name.place(x=20, y=160)
        show_id.place(x=20, y=200)
        show_email.place(x=20, y=240)
        show_phone_number.place(x=20, y=280)

        data_show_name.place(x=170, y=160)
        data_show_id.place(x=170, y=200)
        data_show_email.place(x=170,y=240)
        data_show_phone.place(x=170, y=280)


    def hide_info(event):
        introduction.place(x=7, y=120)

        title.place_forget()
        show_name.place_forget()
        show_id.place_forget()
        show_email.place_forget()
        show_phone_number.place_forget()

        data_show_name.place_forget()
        data_show_id.place_forget()
        data_show_email.place_forget()
        data_show_phone.place_forget()



    profile_photo.bind("<Enter>", user_info)
    profile_photo.bind("<Leave>", hide_info)


#----------------------------------------------------------------MAIN INTERFACE----------------------------------------



origin.mainloop()