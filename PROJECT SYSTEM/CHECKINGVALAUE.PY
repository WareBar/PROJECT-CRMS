from openpyxl import load_workbook
from openpyxl import *


book = load_workbook("CRMSDB.xlsx")
LOG_IN = book["CREDENTIALS"]


for i in range(2, LOG_IN.max_row):
    print(f"{LOG_IN['A'+str(LOG_IN.max_row)].value}")